# Builds the two-mode webgrader report (HTML+PDF via Chrome) and an xlsx.
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# slug,provider,ats,overall,grade,seo,geo,cx,brand,technical,conversion  (None overall = blocked)
REC = [
 ("trsresourcing.com","Applyflow","",82,"B+",90,81,82,78,71,85),
 ("designandbuild.com.au","Shazamme","",81,"B+",84,69,71,93,90,93),
 ("ambition.com.au","Volcanic","",80,"B+",83,59,84,88,79,100),
 ("beaumontpeople.com.au","Shazamme","",80,"B+",78,73,75,83,87,96),
 ("sharpandcarter.com.au","Volcanic","",79,"B+",88,72,80,74,78,80),
 ("marqueestaffing.com","Haley","",79,"B+",82,63,85,83,76,96),
 ("avidtr.com","Staffing Futures","",79,"B+",78,59,90,83,74,96),
 ("mjdrecruitment.com.au","refari","",79,"B+",90,87,77,57,72,88),
 ("doveandhawk.co.uk","Recsites","",77,"B+",69,65,92,97,49,95),
 ("eliasrecruitment.com","refari","",77,"B+",79,77,72,66,76,100),
 ("aspectpersonnel.com.au","Shazamme","",77,"B+",68,59,81,93,93,80),
 ("abbatt.co.uk","Recsites","",74,"B",70,64,92,80,59,73),
 ("dunhillstaff.com","Haley","",72,"B",77,68,68,74,76,65),
 ("primeteampartners.com","Applyflow","",72,"B",66,59,88,74,64,80),
 ("pearlrec.com.au","Jobadder","",71,"B",60,57,92,71,72,71),
 ("emeraldtalent.com","Staffing Futures","",65,"B",67,51,77,60,72,61),
 ("impellam.com","Sourceflow","",64,"C+",52,47,67,90,81,56),
 ("scrhired.com","WordPress","Bullhorn",57,"C+",34,44,81,61,61,76),
 ("caminosearch.co.uk","Sourceflow","",57,"C+",47,47,69,78,74,20),
 ("kofi-group.com","Staffing Futures","",None,"BLOCKED",None,None,None,None,None,None),
]
CAR = [
 ("jobs.sutherlandglobal.com","Shazamme","SmartRecruiters",80,"B+",70,59,79,92,88,96),
 ("careers.tollgroup.com","Shazamme","Workday",78,"B+",64,62,84,89,91,76),
 ("svhm.org.au/careers","elcomCMS","Workday",67,"B",46,52,95,67,78,31),
 ("linfox.com/careers","WordPress","Pageup",62,"C+",45,55,71,71,69,40),
 ("smiggle.com.au/careers","Custom","Pageup",55,"C+",38,56,70,49,62,50),
 ("careers.reece.com/anz","Next.js","Workday",54,"C",52,38,84,33,78,33),
 ("bhp.com/careers","SuccessFactors","SuccessFactors",None,"BLOCKED",None,None,None,None,None,None),
]
PILL=[("seo","SEO"),("geo","GEO/AI"),("cx","CX"),("brand","Brand"),("technical","Tech"),("conversion","Conv")]
WEIGHTS={"recruitment":{"seo":22,"geo":20,"cx":22,"brand":14,"technical":13,"conversion":9},
         "career_site":{"seo":18,"geo":15,"cx":25,"brand":25,"technical":10,"conversion":7}}

def rows(data):
    out=[]
    for slug,prov,ats,ov,gr,se,ge,cx,br,te,co in data:
        out.append(dict(slug=slug,provider=prov,ats=ats,overall=ov,grade=gr,
                        seo=se,geo=ge,cx=cx,brand=br,technical=te,conversion=co,blocked=(ov is None)))
    return out
rec=rows(REC); car=rows(CAR)
def avg(data,key):
    v=[d[key] for d in data if d[key] is not None]; return round(sum(v)/len(v)) if v else 0

SP=os.path.dirname(__file__)
with open(os.path.join(SP,"final_results.csv"),"w",newline="") as f:
    w=csv.writer(f); w.writerow(["mode","site","provider","ats","overall","grade"]+[k for k,_ in PILL])
    for m,data in (("recruitment",rec),("career_site",car)):
        for d in data:
            w.writerow([m,d["slug"],d["provider"],d["ats"],d["overall"],d["grade"]]+[d[k] for k,_ in PILL])

def cls(v): return "" if v is None else ("g" if v>=80 else ("a" if v>=60 else "r"))
def gcls(g):
    if g=="BLOCKED": return "blk"
    return "g" if g.startswith(("A","B+")) else ("a" if g.startswith("B") or g=="C+" else "r")
grec=[d for d in rec if not d['blocked']]; gcar=[d for d in car if not d['blocked']]
rec_avg={k:avg(grec,k) for k,_ in PILL}; car_avg={k:avg(gcar,k) for k,_ in PILL}

def section(title, mode, data, subtitle):
    w=WEIGHTS[mode]; wrow=" · ".join(f"{lbl} {w[k]}%" for k,lbl in PILL)
    grad=[d for d in data if not d["blocked"]]
    head="".join(f"<th>{lbl}<br><span class=wt>{w[k]}%</span></th>" for k,lbl in PILL)
    body=""
    for i,d in enumerate(sorted(data,key=lambda x:(x["overall"] is None,-(x["overall"] or 0))),1):
        shz=d["provider"]=="Shazamme"; prov=d["provider"]+(f" / {d['ats']}" if d["ats"] else "")
        if d["blocked"]:
            cells='<td colspan="6" class="blk">Not gradable — HTTP 403 (bot/WAF block)</td>'
            ov='<td class="blk">—</td>'; gr='<td class="blk">BLOCKED</td>'
        else:
            cells="".join(f'<td class="{cls(d[k])}">{d[k]}</td>' for k,_ in PILL)
            ov=f'<td class="{cls(d["overall"])}"><b>{d["overall"]}</b></td>'; gr=f'<td class="{gcls(d["grade"])}">{d["grade"]}</td>'
        st=" style=font-weight:700" if shz else ""
        body+=f'<tr><td class="l">{i}</td><td class="l"{st}>{d["slug"]}</td><td class="l"{st}>{prov}</td>{ov}{gr}{cells}</tr>'
    av=(f'<tr class="avg"><td class="l" colspan="3">Average (gradable, n={len(grad)})</td>'
        f'<td>{avg(grad,"overall")}</td><td></td>'+"".join(f'<td>{avg(grad,k)}</td>' for k,_ in PILL)+"</tr>")
    return (f'<h2>{title}</h2><p class="mode">Scored in <b>{mode}</b> mode · pillar weighting: {wrow}</p>'
            f'<p class="sub2">{subtitle}</p><table><thead><tr><th class="l">#</th><th class="l">Site</th>'
            f'<th class="l">Provider / ATS</th><th>Overall</th><th>Grade</th>{head}</tr></thead>'
            f'<tbody>{body}{av}</tbody></table>')

HTML=f"""<!doctype html><html><head><meta charset=utf-8><title>Webgrader — Recruitment vs Career</title>
<style>
@page{{size:A4 landscape;margin:11mm 9mm}}*{{box-sizing:border-box}}
body{{font-family:-apple-system,'Helvetica Neue',Arial,sans-serif;color:#15161c;font-size:10.5px;line-height:1.45;margin:0}}
h1{{font-size:21px;color:#1F3864;margin:0 0 2px}}h2{{font-size:14px;color:#1F3864;border-bottom:2px solid #1F3864;padding-bottom:3px;margin:18px 0 4px}}
.sub{{color:#6b7280;margin:0 0 3px}}.mode{{margin:2px 0;font-size:10px;color:#374151}}.sub2{{margin:0 0 6px;color:#6b7280;font-size:10px}}
.wt{{font-weight:400;color:#c9d3e6;font-size:8px}}
table{{border-collapse:collapse;width:100%;margin:3px 0 4px}}th,td{{border:1px solid #d9d9d9;padding:3px 5px;text-align:center}}
th{{background:#1F3864;color:#fff;font-size:9px}}td.l,th.l{{text-align:left}}
tbody tr:nth-child(even) td{{background:#fafbfc}}
.g{{background:#C6EFCE!important}}.a{{background:#FFEB9C!important}}.r{{background:#FFC7CE!important}}.blk{{background:#e5e7eb!important;color:#6b7280}}
.avg td{{font-weight:700;background:#eef1f6!important;border-top:2px solid #1F3864}}
.legend{{font-size:9px;color:#6b7280;margin:2px 0 8px}}.chip{{display:inline-block;width:9px;height:9px;border-radius:2px;margin:0 3px 0 10px;vertical-align:middle}}
ol{{margin:3px 0 3px 16px}}li{{margin:0 0 5px}}.pb{{page-break-before:always}}
.kpi{{display:flex;gap:10px;margin:6px 0}}.card{{border:1px solid #d9d9d9;border-radius:6px;padding:6px 10px;flex:1}}
.card b{{font-size:19px;color:#1F3864}}.card span{{font-size:9px;color:#6b7280}}
.note{{background:#fff8e1;border:1px solid #ffe082;border-radius:5px;padding:7px 9px;font-size:9.5px;color:#5d4037;margin-top:8px}}
footer{{margin-top:10px;font-size:8px;color:#9ca3af;text-align:center}}
</style></head><body>
<h1>Website Grader — Recruitment vs. Career-site modes</h1>
<p class="sub">Graded with the Shazamme Career Site Grader. Each group is scored in its own mode (different pillar weights &amp; checks), so the two tables are reported separately and are <b>not</b> directly comparable to each other.</p>
<p class="legend"><span class="chip g"></span>80–100<span class="chip a"></span>60–79<span class="chip r"></span>&lt;60<span class="chip blk"></span>blocked (403)</p>
<div class="kpi">
 <div class="card"><b>{avg(grec,'overall')}</b><br><span>Recruitment avg (n={len(grec)})</span></div>
 <div class="card"><b>{avg(gcar,'overall')}</b><br><span>Career-site avg (n={len(gcar)})</span></div>
 <div class="card"><b>80 / 78</b><br><span>Shazamme career sites (sutherland / tollgroup) — top of career group</span></div>
 <div class="card"><b>2</b><br><span>Ungradable — kofi-group &amp; BHP hard-block the crawler (403)</span></div>
</div>
{section("1 · Recruitment / staffing sites","recruitment",rec,"Recruitment mode rewards discoverability (SEO+GEO = 42%) and candidate experience — how agencies get found and drive applications.")}
<div class="pb"></div>
{section("2 · Corporate career sites","career_site",car,"Career-site mode rewards employer brand + candidate experience (Brand+CX = 50%) — EVP, culture content and application flow, not job-board SEO.")}
<h2>Analysis</h2>
<ol>
<li><b>The two modes tell different stories — by design.</b> An agency site and a corporate careers site are judged on different things, so they're scored and ranked separately. Recruitment weights SEO+GEO at 42% (get found, rank for "[sector] jobs"); career-site weights Brand+CX at 50% (EVP, culture, application experience). The same URL would score differently in the other mode.</li>
<li><b>Career sites are where Shazamme wins decisively.</b> On the career-site rubric the two Shazamme sites — sutherland <b>80</b> and tollgroup <b>78</b> — are the only B+ grades. Every incumbent-ATS career site trails: svhm 67, linfox 62, smiggle 55, reece 54. The gap is exactly where the rubric weights hardest — <b>Brand</b> (Shazamme 92/89 vs Workday/Pageup 33–71) and <b>Technical</b> (88/91 vs 62–78). Headline pitch: career sites on Workday / Pageup / SuccessFactors underinvest in the two things that matter most for talent attraction.</li>
<li><b>In recruitment, Shazamme is strong but the field is tighter.</b> designandbuild <b>81</b>, ambition/beaumontpeople <b>80</b>, and a B+ cluster (trsresourcing 82, sharpandcarter/marqueestaffing/avidtr 79) sit together. Shazamme's edge is <b>Brand + Technical</b> (designandbuild 93/90, aspectpersonnel 93/93, beaumontpeople 83/87) — consistently top-tier — while pure recruitment-SaaS rivals win on raw SEO (trsresourcing 90, mjdrecruitment 90) but give ground on brand/technical polish.</li>
<li><b>Weakest performers:</b> caminosearch <b>57</b> (SEO 47, conversion 20 — almost no apply/engagement signals), scrhired <b>57</b> (SEO 34 — near-invisible), impellam <b>64</b> (SEO/GEO 52/47 despite strong brand). In the career group, reece <b>54</b> (brand 33, conversion 33) and smiggle <b>55</b> are the clearest rebuild candidates.</li>
<li><b>GEO / AI visibility is the industry-wide gap.</b> It's the lowest pillar in both groups (recruitment avg {rec_avg['geo']}, career avg {car_avg['geo']}) — almost no one is optimised for AI answer engines. First-mover advantage for whoever fixes it.</li>
<li><b>Two sites hard-block automated grading (403):</b> kofi-group and BHP return 403 to the crawler (Cloudflare-class WAF) on both local and production runs. Not a score — but the same block also stops the AI crawlers that increasingly drive candidate discovery.</li>
</ol>
<div class="note"><b>Method:</b> Shazamme Career Site Grader, run per-group in its correct mode via the production service (PageSpeed + headless render enabled). Overall = weighted blend of the six pillars using the mode weights shown in each header. Single run per site — treat ±3–5 as noise. WAF-blocked sites (403) can't be crawled by the tool or by AI crawlers.</div>
<footer>Shazamme Career Site Grader · recruitment &amp; career_site modes · {len(grec)+len(gcar)} of 27 sites graded</footer>
</body></html>"""
open(os.path.join(SP,"report2.html"),"w").write(HTML)

wb=Workbook(); hfill=PatternFill("solid",fgColor="1F3864"); hf=Font(bold=True,color="FFFFFF")
thin=Side(style="thin",color="D9D9D9"); bd=Border(thin,thin,thin,thin)
def fill(v):
    if v is None: return PatternFill("solid",fgColor="E5E7EB")
    return PatternFill("solid",fgColor="C6EFCE" if v>=80 else ("FFEB9C" if v>=60 else "FFC7CE"))
def sheet(ws,data,mode):
    w=WEIGHTS[mode]; hdr=["Site","Provider","ATS","Overall","Grade"]+[f"{lbl} ({w[k]}%)" for k,lbl in PILL]
    ws.append(hdr)
    for c in ws[1]: c.font=hf; c.fill=hfill; c.alignment=Alignment(horizontal="center",wrap_text=True); c.border=bd
    for d in sorted(data,key=lambda x:(x["overall"] is None,-(x["overall"] or 0))):
        row=[d["slug"],d["provider"],d["ats"],d["overall"] if d["overall"] is not None else "—",d["grade"]]+[
            (d[k] if d[k] is not None else "—") for k,_ in PILL]
        ws.append(row); r=ws.max_row
        for ci in range(1,len(hdr)+1):
            ws.cell(r,ci).border=bd; ws.cell(r,ci).alignment=Alignment(horizontal="center" if ci>3 else "left")
        for ci,(k,_) in enumerate(PILL,6): ws.cell(r,ci).fill=fill(d[k])
        ws.cell(r,4).fill=fill(d["overall"])
    for i,wd in enumerate([26,18,15,9,8,10,10,8,9,8,8],1): ws.column_dimensions[get_column_letter(i)].width=wd
    ws.freeze_panes="A2"
sheet(wb.active,rec,"recruitment"); wb.active.title="Recruitment"
sheet(wb.create_sheet("Career sites"),car,"career_site")
wb.save("/Users/rickmare/Downloads/Webgrader - Mode Graded.xlsx")
print("HTML + XLSX written | recruitment avg",avg(grec,'overall'),"| career avg",avg(gcar,'overall'))
