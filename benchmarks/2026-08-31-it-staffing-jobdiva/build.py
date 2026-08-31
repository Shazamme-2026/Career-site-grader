import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SP = os.path.dirname(os.path.abspath(__file__))
PILL = [("seo", "SEO"), ("geo", "GEO/AI"), ("cx", "CX"),
        ("brand", "Brand"), ("technical", "Tech"), ("conversion", "Conv")]
WEIGHTS = {"seo": 22, "geo": 20, "cx": 22, "brand": 14, "technical": 13, "conversion": 9}

ats = {}
vpath = os.path.join(SP, "verdict.csv")
if os.path.exists(vpath):
    for r in csv.DictReader(open(vpath)):
        ats[r["site"]] = r["ats"]

rows = []
with open(os.path.join(SP, "results.csv")) as f:
    for r in csv.DictReader(f):
        def num(v):
            return int(v) if str(v).strip().isdigit() else None
        rows.append({"site": r["site"], "platform": r.get("platform", ""),
                     "ats": ats.get(r["site"], "?"),
                     "overall": num(r["overall"]), "grade": r["grade"],
                     **{k: num(r[k]) for k, _ in PILL},
                     "blocked": not str(r["overall"]).strip().isdigit()})

grad = [d for d in rows if not d["blocked"]]
def avg(key):
    v = [d[key] for d in grad if d[key] is not None]
    return round(sum(v) / len(v)) if v else 0

rows.sort(key=lambda x: (x["overall"] is None, -(x["overall"] or 0)))

wb = Workbook()
ws = wb.active; ws.title = "Recruitment"
hfill = PatternFill("solid", fgColor="1F3864"); hf = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="D9D9D9"); bd = Border(thin, thin, thin, thin)
def fill(v):
    if v is None:
        return PatternFill("solid", fgColor="E5E7EB")
    return PatternFill("solid", fgColor="C6EFCE" if v >= 80 else ("FFEB9C" if v >= 60 else "FFC7CE"))

JDFILL = PatternFill("solid", fgColor="C6EFCE")
OTHERFILL = PatternFill("solid", fgColor="FFC7CE")
UNKFILL = PatternFill("solid", fgColor="FFEB9C")
def atsfill(v):
    if v == "JobDiva": return JDFILL
    if v in ("?", "—"): return UNKFILL
    return OTHERFILL

hdr = ["#", "Site", "ATS (apply page)", "Platform", "Overall", "Grade"] + [f"{lbl} ({WEIGHTS[k]}%)" for k, lbl in PILL]
ws.append(hdr)
for c in ws[1]:
    c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = bd
for i, d in enumerate(rows, 1):
    row = [i, d["site"], d["ats"], d["platform"],
           d["overall"] if d["overall"] is not None else "—", d["grade"]] + [
           (d[k] if d[k] is not None else "—") for k, _ in PILL]
    ws.append(row); r = ws.max_row
    for ci in range(1, len(hdr) + 1):
        ws.cell(r, ci).border = bd
        ws.cell(r, ci).alignment = Alignment(horizontal="center" if ci not in (2, 3, 4) else "left")
    for ci, (k, _) in enumerate(PILL, 7):
        ws.cell(r, ci).fill = fill(d[k])
    ws.cell(r, 5).fill = fill(d["overall"])
    ws.cell(r, 3).fill = atsfill(d["ats"])
ws.append(["", f"Average (gradable, n={len(grad)})", "", "", avg("overall"), ""] + [avg(k) for k, _ in PILL])
ar = ws.max_row
for ci in range(1, len(hdr) + 1):
    c = ws.cell(ar, ci); c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor="EEF1F6"); c.border = bd
    c.alignment = Alignment(horizontal="center" if ci not in (2, 3, 4) else "left")
for i, wd in enumerate([4, 24, 16, 15, 9, 8, 10, 10, 8, 9, 8, 8], 1):
    ws.column_dimensions[get_column_letter(i)].width = wd
ws.freeze_panes = "A2"

out = "/Users/rickmare/Downloads/Webgrader - IT Staffing & Consulting.xlsx"
wb.save(out)
print(f"XLSX -> {out}")
print(f"graded {len(grad)}/{len(rows)} | avg overall {avg('overall')} | "
      f"blocked: {', '.join(d['site'] for d in rows if d['blocked']) or 'none'}")
